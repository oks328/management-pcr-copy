from flask import render_template, request, redirect, session, url_for, flash, send_file
from app import app
from .auth import hash_pass, verify_pass
from .models import get_db_connection, get_user_by_email, register_user, get_pending_users, update_user_status, open_new_term, get_all_terms, add_master_indicator, get_master_indicators, reset_user_password, import_previous_term_indicators, edit_master_indicator, delete_master_indicator, get_cascaded_quotas, cascade_institutional_targets
import mysql.connector, time
from functools import wraps
import os, subprocess, datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'Admin':
            flash("Unauthorized access. Admin privileges required.", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def dean_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'Dean':
            flash("Unauthorized access. Dean privileges required.", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def login():
    return render_template('login.html')

@app.route('/authenticate', methods=['POST'])
def authenticate():
    email = request.form.get('email', '').lower().strip()
    password = request.form.get('password', '')
    
    conn = get_db_connection()
    if not conn:
        flash("Database connection failed.", "danger")
        return redirect(url_for('login'))
        
    cursor = conn.cursor()
    result = get_user_by_email(cursor, email)
    conn.close()
    
    if not result:
        time.sleep(0.5) # prevent timing attacks
        flash("Invalid Credentials.", "danger")
        return redirect(url_for('login'))
    
    row = result[0]
    emp_id, stored_hash, system_role, verification_status, first_name, last_name = row
    
    if verify_pass(password, stored_hash):
        if verification_status == 'PENDING':
            flash("Account awaiting Admin approval.", "warning")
            return redirect(url_for('login'))
        elif verification_status == 'REJECTED':
            flash("Account has been rejected.", "danger")
            return redirect(url_for('login'))
            
        session['emp_id'] = emp_id
        session['role'] = system_role
        session['name'] = first_name
        
        if system_role == 'Admin': return redirect(url_for('admin_dashboard'))
        elif system_role == 'Dean': return redirect(url_for('dean_dashboard'))
        elif system_role == 'Manager': return redirect(url_for('manager_dashboard'))
        elif system_role == 'Faculty': return redirect(url_for('faculty_dashboard'))
        elif system_role == 'Designated': return redirect(url_for('designated_dashboard'))
        
        return redirect(url_for('login'))
    else:
        time.sleep(0.5)
        flash("Invalid Credentials.", "danger")
        return redirect(url_for('login'))
        
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        emp_id = request.form.get('emp_id', '').strip()
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        college = request.form.get('college', '').strip()
        academic_rank = request.form.get('academic_rank', '').strip()
        employment_status = request.form.get('employment_status', '').strip()
        assigned_program = request.form.get('assigned_program', '').strip()
        designation = request.form.get('designation', '').strip()
        system_role = request.form.get('system_role', 'Faculty').strip()
        email = request.form.get('email', '').lower().strip()
        password = request.form.get('password', '')
        
        if not designation:
            designation = 'None'
            
        if not all([emp_id, first_name, last_name, college, academic_rank, employment_status, assigned_program, email, password]):
            flash("All fields required.", "danger")
            return redirect(url_for('register'))
        
        hashed_pw = hash_pass(password)
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "danger")
            return redirect(url_for('register'))
            
        try:
            cursor = conn.cursor()
            register_user(conn, cursor, emp_id, first_name, last_name, college, academic_rank, employment_status, assigned_program, designation, email, hashed_pw, system_role)
            cursor.close()
            conn.close()
            flash("Registration successful. Awaiting Admin approval.", "success")
            return redirect(url_for('login'))
        except mysql.connector.Error as e:
            if conn.is_connected():
                conn.rollback()
                conn.close()
            flash("Registration failed. Employee ID or Email may already exist.", "danger")
            return redirect(url_for('register'))
        
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for('login'))

@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    if not conn:
        flash("Database connection failed.", "danger")
        return render_template('admin_dashboard.html')
        
    cursor = conn.cursor()
    
    pending_users = get_pending_users(cursor)
    terms = get_all_terms(cursor)
    
    # Calculate stats
    cursor.execute("SELECT COUNT(*) FROM tbl_auth_credentials WHERE verification_status = 'APPROVED'")
    active_users = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM tbl_evidence_repo")
    evidence_count = cursor.fetchone()[0]
    
    # Find active term details
    cursor.execute("SELECT term_id, academic_year, semester, deadline_date FROM tbl_academic_terms WHERE is_active = TRUE")
    active_term = cursor.fetchone()
    days_remaining = 0
    if active_term and active_term[3]:
        delta = active_term[3] - datetime.date.today()
        days_remaining = delta.days if delta.days > 0 else 0
        
    indicators = []
    if active_term:
        indicators = get_master_indicators(cursor, active_term[0])
        
    stats = {
        'active_users': active_users,
        'evidence_count': evidence_count,
        'days_remaining': days_remaining
    }
    
    conn.close()
    return render_template('admin_dashboard.html', pending_users=pending_users, terms=terms, indicators=indicators, stats=stats, active_term=active_term)

@app.route('/admin/verify_user', methods=['POST'])
@admin_required
def admin_verify_user():
    emp_id = request.form.get('emp_id')
    action = request.form.get('action') # 'approve' or 'reject'
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        update_user_status(conn, cursor, emp_id, action)
        flash(f"User {action.upper()}D successfully.", "success")
    except Exception as e:
        flash(f"Error processing action: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/open_term', methods=['POST'])
@admin_required
def admin_open_term():
    academic_year = request.form.get('academic_year')
    semester = request.form.get('semester')
    deadline_date = request.form.get('deadline_date')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        open_new_term(conn, cursor, academic_year, semester, deadline_date)
        flash("New Academic Term opened successfully.", "success")
    except Exception as e:
        flash(f"Error opening term: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/add_indicator', methods=['POST'])
@admin_required
def admin_add_indicator():
    term_id = request.form.get('term_id')
    category_name = request.form.get('category_name')
    description = request.form.get('description')
    efficiency_type = request.form.get('efficiency_type')
    
    if not term_id:
        flash("Action denied. No active Academic Term found.", "danger")
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        add_master_indicator(conn, cursor, category_name, description, efficiency_type, term_id)
        flash("Master Indicator added successfully.", "success")
    except Exception as e:
        flash(f"Error adding indicator: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/indicators/import', methods=['POST'])
@admin_required
def admin_import_indicators():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT term_id FROM tbl_academic_terms WHERE is_active = TRUE")
        active_term = cursor.fetchone()
        if not active_term:
            flash("Action denied. No active Academic Term found.", "danger")
            return redirect(url_for('admin_dashboard'))
            
        success, msg = import_previous_term_indicators(conn, cursor, active_term[0])
        if success:
            flash(msg, "success")
        else:
            flash(msg, "warning")
    except Exception as e:
        flash(f"Error importing indicators: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/indicators/edit', methods=['POST'])
@admin_required
def admin_edit_indicator():
    indicator_id = request.form.get('indicator_id')
    category_name = request.form.get('category_name')
    description = request.form.get('description')
    efficiency_type = request.form.get('efficiency_type')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        edit_master_indicator(conn, cursor, indicator_id, category_name, description, efficiency_type)
        flash("Master Indicator updated successfully.", "success")
    except Exception as e:
        flash(f"Error updating indicator: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/indicators/delete', methods=['POST'])
@admin_required
def admin_delete_indicator():
    indicator_id = request.form.get('indicator_id')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        delete_master_indicator(conn, cursor, indicator_id)
        flash("Master Indicator deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting indicator: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/reset_password', methods=['POST'])
@admin_required
def admin_reset_password():
    emp_id = request.form.get('emp_id')
    new_password = request.form.get('new_password')
    
    if not new_password:
        flash("Password cannot be empty.", "danger")
        return redirect(url_for('admin_dashboard'))
        
    hashed_pw = hash_pass(new_password)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        reset_user_password(conn, cursor, emp_id, hashed_pw)
        flash("Password reset successfully.", "success")
    except Exception as e:
        flash(f"Error resetting password: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/backup_db')
@admin_required
def admin_backup_db():
    backup_dir = os.path.join(app.root_path, 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
        
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"ipcr_db_backup_{timestamp}.sql"
    filepath = os.path.join(backup_dir, filename)
    
    try:
        cmd = ['mysqldump', '-u', 'root', 'ipcr_db']
        with open(filepath, 'w') as f:
            subprocess.run(cmd, stdout=f, stderr=subprocess.PIPE, check=True)
        flash(f"Database backed up successfully to: /backups/{filename}", "success")
    except FileNotFoundError:
        try:
            xampp_cmd = [r'C:\xampp\mysql\bin\mysqldump.exe', '-u', 'root', 'ipcr_db']
            with open(filepath, 'w') as f:
                subprocess.run(xampp_cmd, stdout=f, stderr=subprocess.PIPE, check=True)
            flash(f"Database backed up successfully to: /backups/{filename}", "success")
        except Exception as fallback_e:
            flash(f"Backup failed. Could not find mysqldump: {fallback_e}", "danger")
    except Exception as e:
        flash(f"Backup failed: {e}", "danger")
        
    return redirect(url_for('admin_dashboard'))

@app.route('/faculty')
def faculty_dashboard(): return render_template('faculty_dashboard.html')

@app.route('/dean')
@dean_required
def dean_dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT term_id, academic_year, semester FROM tbl_academic_terms WHERE is_active = TRUE")
    active_term = cursor.fetchone()
    
    indicators = []
    grouped_indicators = {}
    quotas = {}
    if active_term:
        term_id = active_term[0]
        indicators = get_master_indicators(cursor, term_id)
        for row in indicators:
            cat_name = row[1]
            if cat_name not in grouped_indicators:
                grouped_indicators[cat_name] = []
            grouped_indicators[cat_name].append({
                'indicator_id': row[0],
                'indicator_description': row[2],
                'efficiency_type': row[3]
            })
        quotas = get_cascaded_quotas(cursor, term_id)
        
    conn.close()
    return render_template('dean_dashboard.html', active_term=active_term, indicators=indicators, grouped_indicators=grouped_indicators, quotas=quotas)

@app.route('/dean/review_targets', methods=['POST'])
@dean_required
def dean_review_targets():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT term_id, academic_year, semester FROM tbl_academic_terms WHERE is_active = TRUE")
    active_term = cursor.fetchone()
    
    if not active_term:
        conn.close()
        flash("No active term found.", "warning")
        return redirect(url_for('dean_dashboard'))
        
    term_id = active_term[0]
    indicators = get_master_indicators(cursor, term_id)
    indicator_dict = {ind[0]: ind for ind in indicators}
    
    review_list = []
    for key, value in request.form.items():
        if key.startswith('quota_'):
            val = str(value).strip()
            if val and val != "0":
                parts = key.split('_', 2)
                if len(parts) == 3:
                    indicator_id = int(parts[1])
                    role = parts[2]
                    
                    if indicator_id in indicator_dict:
                        ind = indicator_dict[indicator_id]
                        review_list.append({
                            'indicator_id': indicator_id,
                            'role': role,
                            'value': int(val),
                            'category': ind[1],
                            'description': ind[2]
                        })
                        
    conn.close()
    
    if not review_list:
        flash("No valid quotas entered. Please ensure values are greater than 0.", "warning")
        return redirect(url_for('dean_dashboard'))
        
    return render_template('dean_review_targets.html', active_term=active_term, review_list=review_list)

@app.route('/dean/cascade_targets', methods=['POST'])
@dean_required
def dean_cascade_targets():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT term_id FROM tbl_academic_terms WHERE is_active = TRUE")
        active_term = cursor.fetchone()
        
        if not active_term:
            flash("No active term found to assign targets for.", "warning")
            return redirect(url_for('dean_dashboard'))
            
        term_id = active_term[0]
        targets_list = []
        
        for key, value in request.form.items():
            if key.startswith('quota_') and value:
                val = str(value).strip()
                if val and val != "0":
                    parts = key.split('_', 2)
                    if len(parts) == 3:
                        indicator_id = int(parts[1])
                        role = parts[2]
                        targets_list.append((term_id, indicator_id, int(val), role))
                        
        cascade_institutional_targets(conn, cursor, term_id, targets_list)
        flash("Institutional targets successfully reviewed and cascaded to programs.", "success")
    except Exception as e:
        flash(f"Error cascading targets: {str(e)}", "danger")
    finally:
        conn.close()
        
    return redirect(url_for('dean_dashboard'))

@app.route('/dean/export_dpcr')
@dean_required
def dean_export_dpcr():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT term_id, academic_year, semester FROM tbl_academic_terms WHERE is_active = TRUE")
    active_term = cursor.fetchone()
    if not active_term:
        conn.close()
        flash("No active term found to export.", "warning")
        return redirect(url_for('dean_dashboard'))
        
    term_id = active_term[0]
    indicators = get_master_indicators(cursor, term_id)
    quotas = get_cascaded_quotas(cursor, term_id)
    conn.close()
    
    grouped_indicators = {}
    for row in indicators:
        cat_name = row[1]
        if cat_name not in grouped_indicators:
            grouped_indicators[cat_name] = []
            
        qu = {
            'WST': quotas.get((row[0], 'WST'), 0),
            'DST': quotas.get((row[0], 'DST'), 0),
            'NST': quotas.get((row[0], 'NST'), 0),
            'BSDS': quotas.get((row[0], 'BSDS'), 0),
            'RET': quotas.get((row[0], 'RET'), 0),
            'CICT_Shared': quotas.get((row[0], 'CICT_Shared'), 0)
        }
        
        grouped_indicators[cat_name].append({
            'indicator_description': row[2],
            'quotas': qu
        })
        
    template_path = os.path.join(app.root_path, 'dpcr_template.xlsx')
    if not os.path.exists(template_path):
        flash("DPCR Template not found on server.", "danger")
        return redirect(url_for('dean_dashboard'))
        
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    insert_row = 12
    for r in range(1, 100):
        cell_a = sheet.cell(row=r, column=1).value
        cell_b = sheet.cell(row=r, column=2).value
        
        if (cell_a and "Total Overall Rating" in str(cell_a)) or \
           (cell_b and "Total Overall Rating" in str(cell_b)):
            insert_row = r
            break
            
    for cat_name, items in grouped_indicators.items():
        sheet.insert_rows(insert_row, amount=1)
        sheet.merge_cells(start_row=insert_row, start_column=1, end_row=insert_row, end_column=18)
        
        cat_cell = sheet.cell(row=insert_row, column=1)
        cat_cell.value = cat_name
        cat_cell.font = Font(bold=True)
        cat_cell.fill = header_fill
        cat_cell.alignment = left_align
        cat_cell.border = thin_border
        
        insert_row += 1
        
        for item in items:
            sheet.insert_rows(insert_row, amount=1)
            
            # Explicitly merge Columns A-D to keep the MFO/PAP column blank but structured
            sheet.merge_cells(start_row=insert_row, start_column=1, end_row=insert_row, end_column=4)
            
            # Explicitly merge Columns E-G for the Success Indicator text!
            sheet.merge_cells(start_row=insert_row, start_column=5, end_row=insert_row, end_column=7)
            
            # Write indicator description to Column E (Index 5)
            desc_cell = sheet.cell(row=insert_row, column=5)
            desc_cell.value = item['indicator_description']
            desc_cell.alignment = left_align
            
            # Apply borders to the merged sections
            for col in range(1, 8):
                sheet.cell(row=insert_row, column=col).border = thin_border
                
            q = item['quotas']
            cict = q.get('CICT_Shared', 0)
            
            # The Column Logic (Shared vs Program)
            if cict and int(cict) > 0:
                # Merge Columns I through M (9 to 13)
                sheet.merge_cells(start_row=insert_row, start_column=9, end_row=insert_row, end_column=13)
                val_cell = sheet.cell(row=insert_row, column=9)
                val_cell.value = int(cict)
                val_cell.alignment = center_align
                
                # Apply borders across the merged shared columns
                for col in range(9, 14):
                    sheet.cell(row=insert_row, column=col).border = thin_border
            else:
                mapping = {9: 'DST', 10: 'WST', 11: 'NST', 12: 'BSDS', 13: 'RET'}
                for col_idx, key in mapping.items():
                    val = q.get(key, 0)
                    if val and int(val) > 0:
                        c = sheet.cell(row=insert_row, column=col_idx)
                        c.value = int(val)
                        c.alignment = center_align
                        c.border = thin_border
                        
            insert_row += 1
            
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Generated_DPCR_{active_term[1]}_{active_term[2]}.xlsx"
    return send_file(output, download_name=filename.replace(' ', '_'), as_attachment=True)

@app.route('/manager')
def manager_dashboard(): return render_template('manager_dashboard.html')

@app.route('/designated')
def designated_dashboard(): return render_template('designated_dashboard.html')